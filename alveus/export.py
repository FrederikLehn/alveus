import os
from datetime import datetime
import numpy as np
import openpyxl as xl

from profile_ import Profile
from variable_mgr import Date


def ExportWriter(writer):
    try:
        writer.save()
    except PermissionError:
        print('Unable to export, the file is in use')


def WriteProfile(ws, profile, variables, entity_names, phaser=False):
    n = profile.time().size

    if phaser:
        headers = ['Date', 'Oil rate (stb/day)', 'Total Gas Rate (kscf/day)', 'Water Rate (stb/day)',
                   'Lift Gas Rate (kscf/day)', 'Gas Injection Rate (kscf/day)', 'Water Injection Rate(stb/day)']

        unit_conversions = np.full(len(variables), 1.e3)

    else:
        headers = ['{} [{}]'.format(v.GetMenuLabel(), v.GetUnit()) for v in variables]
        unit_conversions = np.full(len(variables), 1.)

    # write headers
    for j, _ in enumerate(variables):
        ws.cell(row=1, column=j+2).value = headers[j]

    # write name
    ws.cell(row=1, column=1).value = 'EntityName'
    for i in range(n):
        ws.cell(row=i+2, column=1).value = entity_names[i]

    # write variables
    for j, variable in enumerate(variables):
        array = profile.Get(variable.GetId())

        if isinstance(variable, Date):
            array = array.astype(datetime)
        else:
            array *= unit_conversions[j]

        for i in range(n):
            ws.cell(row=i+2, column=j + 2).value = array[i]


def GetPath(directory, name):
    # pre-append the date
    file_name = datetime.today().strftime('%Y-%m-%d') + '_' + name

    path = os.path.join(directory, file_name)
    if path[-5:] != '.xlsx':
        path += '.xlsx'

    return path


def ToExcel(cases, items, simulations, variables, directory, dateline=None, phaser=False):
    # items is a {{[]}} list with items {file_name: {sheet_name: [entities]}}

    # write to excel
    for case in cases:

        for simulation in simulations:

            for file_name, sheets in items.items():

                wb = xl.Workbook()

                for sheet_name, entities in sheets.items():

                    names = []
                    profiles = Profile()

                    for entity in entities:

                        profile = entity.GetSimulationResult(simulation).GetProfile(idx=case)

                        if dateline is not None:
                            profile = profile.resample(dateline)

                        profiles.stack(profile)
                        names += [entity.GetName()] * profile.time().size

                    # write to sheet
                    ws = wb.create_sheet(title=sheet_name)
                    WriteProfile(ws, profiles, variables, names, phaser=phaser)

                # delete default sheet
                wb.remove_sheet(wb.get_sheet_by_name('Sheet'))

                case_name = 'LOW' if case == 0 else 'MID' if case == 1 else 'HIGH'
                path = GetPath(directory, '{}_{}_{}'.format(simulation.GetName(), file_name, case_name))
                wb.save(path)
